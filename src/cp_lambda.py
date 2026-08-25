#!/usr/bin/env python3
"""
cp_lambda.py — turn a blade sweep into Cp(λ), once rotor speed is available.

    # rotor speed recovered from the generator's three phases
    python src/cp_lambda.py --sweep logs/sweep_v1_Ra20_summary.csv \
           --radius 0.30 --poles 12 --daq reference/data/03162026_sec_backup.xlsx

    # or from a column already in the points CSV
    python src/cp_lambda.py --sweep logs/sweep_v1_Ra20_points.csv --radius 0.30

    # or explore what a radius/pole guess would imply, with no rotor data
    python src/cp_lambda.py --sweep logs/sweep_v1_Ra20_summary.csv \
           --radius 0.30 --assume-lambda 4.0

═══════════════════════════════════════════════════════════════════════════
WHAT IS AND IS NOT COMPUTED
═══════════════════════════════════════════════════════════════════════════
    λ    = ω R / v            tip-speed ratio
    Cp_elec = P / (½ ρ A v³)  ELECTRICAL power coefficient

    A = 2·R·H   for a VERTICAL-axis rotor — it sweeps a cylinder   [default]
    A = π·R²    for a horizontal-axis propeller — it sweeps a disc

The Aerolab rotor is a three-blade VAWT, so its swept area is a rectangle,
not a disc. The two differ by 1.54× here. Getting it wrong scales every Cp by
that factor and nothing in the numbers looks wrong.

**Cp_elec is not Cp.** It is Cp_aero × η_gen × η_rect: everything the rotor
extracted, minus what the generator and rectifier lost. It is the honest
quantity to report from this rig because P is measured at the load terminals
and nothing here measures shaft torque.

Do not compare Cp_elec to the Betz limit. A rotor at Betz with a 40%
efficient generator reads 0.24, and the number means nothing without the
efficiency chain attached to it.

═══════════════════════════════════════════════════════════════════════════
THE TWO NUMBERS THIS NEEDS
═══════════════════════════════════════════════════════════════════════════
**Tip radius**, from the AXIS OF ROTATION — not blade length. λ scales
linearly with it and Cp as 1/R², so a 10% error in radius is a 21% error in
Cp. There is no default and there will not be one.

**Rotor speed.** Three routes, in descending order of directness:

  1. `--rpm-column` — a column already in the sweep CSV. Best.
  2. `--daq` + `--poles` — recovered from the generator's three phases via a
     Clarke transform. The March capture showed ch3/4/6 mutually correlated
     at −0.50 = cos(120°), so the phases are already on the DAQ; only the
     pole count is missing. See `docs/08_march_daq.md`.
  3. `--assume-lambda` — assume a constant tip-speed ratio and back rotor
     speed out of it. This is EXPLORATORY ONLY: it assumes the answer to the
     question Cp(λ) exists to ask, so every point lands at the λ you assumed.
     Useful for sizing an axis, never for a result.
"""

from __future__ import annotations

import argparse
import csv
import math
import sys
from pathlib import Path

import numpy as np

sys.path.insert(0, str(Path(__file__).resolve().parent))

RHO_DEFAULT = 1.225


# ═══════════════════════════════════════════════════════════════════════════
# ROTOR SPEED FROM THE GENERATOR'S THREE PHASES
# ═══════════════════════════════════════════════════════════════════════════

def clarke_frequency(a, b, c, fs, smooth=201):
    """
    Instantaneous electrical frequency from a three-phase set.

    The Clarke transform turns three 120°-separated signals into one rotating
    vector, so the electrical angle is just its argument and the frequency is
    the derivative. That beats zero-crossing detection on every count: it uses
    all three channels, it is immune to amplitude imbalance, and it gives a
    value at every sample rather than twice a cycle.

    Returns |f| in Hz, median-filtered.
    """
    from scipy.signal import medfilt
    a, b, c = np.asarray(a, float), np.asarray(b, float), np.asarray(c, float)
    alpha = (2 * a - b - c) / 3.0
    beta = (b - c) / math.sqrt(3.0)
    phase = np.unwrap(np.arctan2(beta, alpha))
    f = np.gradient(phase, 1.0 / fs) / (2 * math.pi)
    k = smooth if smooth % 2 else smooth + 1
    return np.abs(medfilt(f, k))


def rotor_rpm_from_phases(f_elec_hz, poles):
    """
    Electrical Hz → rotor rpm.

    `poles` is the number of MAGNETIC POLES, not pole pairs. A 12-pole machine
    turns 6 electrical cycles per revolution, so rpm = 60 f / (poles/2).
    Getting this factor wrong scales every λ by the same ratio and every Cp
    with it — it is the single easiest way to produce a plausible wrong curve.
    """
    if poles < 2 or poles % 2:
        raise ValueError("poles must be an even number >= 2 (magnetic poles, "
                         "not pole pairs)")
    return 60.0 * np.asarray(f_elec_hz, float) / (poles / 2.0)


def load_daq_phases(path, cols=(2, 3, 5), fs=None):
    """Pull three phase columns out of the DAQ export."""
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    t = np.array([float(r[0]) for r in rows])
    volts = np.array([[float(x) if x is not None else np.nan
                       for x in r[3:9]] for r in rows]).T
    if fs is None:
        fs = 1.0 / float(np.median(np.diff(t)))
    return t, volts[list(cols)], fs


# ═══════════════════════════════════════════════════════════════════════════

def swept_area(radius_m, hub_m=0.0, rotor="vawt", height_m=None):
    """
    Swept area — and the formula depends on the ROTOR TYPE, not a detail.

    **VAWT** (vertical axis, H-rotor / Darrieus): the blades sweep a CYLINDER,
    and the frontal area the wind sees is its rectangular projection:

        A = 2 · R · H

    **HAWT** (horizontal axis, propeller): the blades sweep a DISC:

        A = π · (R² − hub²)

    On this rig the two differ by 1.54×, so using the wrong one puts every
    Cp out by that factor while every number still looks entirely plausible.
    The Aerolab rotor is a three-blade VAWT — hence the default.
    """
    if rotor == "hawt":
        return math.pi * (radius_m ** 2 - hub_m ** 2)
    if not height_m:
        raise ValueError(
            "a VAWT sweeps a cylinder, so its area needs BLADE HEIGHT as well "
            "as radius: A = 2*R*H. Pass --height, or --rotor hawt if this is "
            "a propeller.")
    return 2.0 * radius_m * height_m


def air_density(temp_c=None, pressure_pa=None):
    if temp_c is None or pressure_pa is None:
        return RHO_DEFAULT
    return pressure_pa / (287.058 * (temp_c + 273.15))


def compute(points, radius_m, hub_m=0.0, rho=RHO_DEFAULT, area=None):
    """
    points: iterable of dicts with mps, p_w, and rpm (rotor).
    Returns the same rows with lam and cp_elec added.
    """
    A = area if area is not None else swept_area(radius_m, hub_m, 'hawt')
    out = []
    for p in points:
        v, P, n = p.get("mps"), p.get("p_w"), p.get("rpm")
        if not v or v <= 0 or n is None:
            continue
        omega = 2 * math.pi * float(n) / 60.0
        avail = 0.5 * rho * A * v ** 3
        out.append({**p,
                    "omega": omega,
                    "lam": omega * radius_m / v,
                    "avail_w": avail,
                    "cp_elec": (P / avail) if avail > 0 else float("nan")})
    return out


def read_sweep(path):
    rows = []
    with open(path) as f:
        body = [l for l in f if not l.startswith("#")]
    for r in csv.DictReader(body):
        try:
            v = float(r.get("wind_mps") or r.get("mps") or 0)
            p = float(r.get("p_max_fit_w") or r.get("p_max_w")
                      or r.get("watts") or 0)
        except (TypeError, ValueError):
            continue
        row = {"mps": v, "p_w": p,
               "fan_rpm": float(r.get("fan_rpm_cmd") or r.get("fan_rpm") or 0)}
        for k in ("rotor_rpm", "rpm"):
            if r.get(k):
                try:
                    row["rpm"] = float(r[k])
                except ValueError:
                    pass
        rows.append(row)
    return rows


# ═══════════════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(
        description="Cp(lambda) from a blade sweep",
        epilog="Reports Cp_elec, not Cp. Do not compare it to the Betz limit.")
    p.add_argument("--sweep", required=True, help="a *_summary.csv or *_points.csv")
    p.add_argument("--radius", type=float, required=True,
                   help="tip radius in METRES, from the axis of rotation — "
                        "NOT blade length. No default: a 10%% error here is a "
                        "21%% error in Cp.")
    p.add_argument("--hub", type=float, default=0.0,
                   help="hub radius, m — HAWT only")
    p.add_argument("--rotor", choices=["vawt", "hawt"], default="vawt",
                   help="vawt (default, sweeps a cylinder: A = 2RH) or hawt "
                        "(sweeps a disc: A = pi R^2). This is not cosmetic — "
                        "the two differ by 1.54x on this rig.")
    p.add_argument("--height", type=float, default=None,
                   help="blade height/span in METRES. Required for a VAWT.")
    p.add_argument("--temp", type=float, default=None, help="air temp °C")
    p.add_argument("--pressure", type=float, default=None, help="Pa")

    g = p.add_argument_group("rotor speed — pick one")
    g.add_argument("--rpm-column", dest="rpm_column",
                   help="a rotor-rpm column already in the CSV")
    g.add_argument("--daq", help="DAQ export carrying the generator phases")
    g.add_argument("--poles", type=int,
                   help="MAGNETIC POLES of the generator (not pole pairs)")
    g.add_argument("--phase-cols", default="2,3,5", dest="phase_cols",
                   help="0-based voltage columns of the three phases "
                        "(default 2,3,5 = ch3/ch4/ch6)")
    g.add_argument("--assume-lambda", type=float, default=None,
                   dest="assume_lambda",
                   help="EXPLORATORY: assume a constant tip-speed ratio. This "
                        "assumes the answer to the question Cp(lambda) exists "
                        "to ask.")
    p.add_argument("--csv", default=None, help="write the result here")
    a = p.parse_args()

    rows = read_sweep(a.sweep)
    if not rows:
        raise SystemExit(f"no usable rows in {a.sweep}")
    rho = air_density(a.temp, a.pressure)

    # ── rotor speed ─────────────────────────────────────────────────────
    src = None
    if a.rpm_column:
        src = f"column '{a.rpm_column}'"
        if not all("rpm" in r for r in rows):
            raise SystemExit(f"{a.sweep} has no rotor-rpm column")
    elif a.daq:
        if not a.poles:
            raise SystemExit("--daq needs --poles (magnetic poles, not pairs)")
        cols = tuple(int(x) for x in a.phase_cols.split(","))
        print(f"  reading generator phases from {Path(a.daq).name}…")
        t, ph, fs = load_daq_phases(a.daq, cols)
        f_elec = clarke_frequency(ph[0], ph[1], ph[2], fs)
        rpm = rotor_rpm_from_phases(f_elec, a.poles)
        keep = t > 30                      # drop the start-up transient
        print(f"    {fs:.0f} Hz, {t[-1]:.0f} s · electrical "
              f"{np.percentile(f_elec[keep],2):.2f}–"
              f"{np.percentile(f_elec[keep],98):.2f} Hz")
        print(f"    → rotor {np.percentile(rpm[keep],2):.0f}–"
              f"{np.percentile(rpm[keep],98):.0f} rpm at {a.poles} poles")
        print(f"\n  ⚠ The DAQ capture and this sweep are DIFFERENT RUNS. They "
              f"cannot be\n    joined point by point without a shared time "
              f"base — see docs/05_integration.md.\n    Reporting the rotor "
              f"range only; per-point Cp needs rotor rpm IN the sweep.\n")
        return 0
    elif a.assume_lambda:
        src = f"ASSUMED constant lambda = {a.assume_lambda}"
        for r in rows:
            r["rpm"] = a.assume_lambda * r["mps"] / a.radius * 60 / (2 * math.pi)
    else:
        raise SystemExit(
            "\n  No rotor speed. Give one of:\n"
            "    --rpm-column <name>     a column already in the CSV\n"
            "    --daq <file> --poles N  recovered from the generator phases\n"
            "    --assume-lambda X       exploratory only\n\n"
            "  Rotor speed is the measurement that separates rotor "
            "aerodynamics\n  from generator matching. Without it this is "
            "P_max(v), not Cp(lambda).\n")

    A_ = swept_area(a.radius, a.hub, a.rotor, a.height)
    res = compute(rows, a.radius, a.hub, rho, A_)
    if not res:
        raise SystemExit("nothing computable")

    A = swept_area(a.radius, a.hub, a.rotor, a.height)
    geom = (f"{a.rotor.upper()}  R = {a.radius:.4f} m" +
            (f", H = {a.height:.4f} m" if a.rotor == 'vawt'
             else f", hub {a.hub:.4f} m"))
    print(f"\n  {geom} → A = {A:.5f} m²"
          f"   ρ = {rho:.4f} kg/m³")
    print(f"  rotor speed: {src}\n")
    print(f"  {'m/s':>6} {'rotor rpm':>10} {'lambda':>8} {'P (W)':>9} "
          f"{'avail (W)':>10} {'Cp_elec':>9}")
    print(f"  {'-'*6} {'-'*10} {'-'*8} {'-'*9} {'-'*10} {'-'*9}")
    for r in res:
        print(f"  {r['mps']:6.1f} {r['rpm']:10.0f} {r['lam']:8.2f} "
              f"{r['p_w']:9.4f} {r['avail_w']:10.2f} {r['cp_elec']:9.5f}")

    cps = [r["cp_elec"] for r in res]
    best = res[int(np.argmax(cps))]
    print(f"\n  peak Cp_elec = {max(cps):.5f} at λ = {best['lam']:.2f}, "
          f"{best['mps']:.1f} m/s")
    print(f"  ── Cp_elec is Cp_aero × η_gen × η_rect. Do NOT compare it to "
          f"the Betz limit.")
    if a.assume_lambda:
        print(f"\n  ⚠ λ was ASSUMED constant, so every point sits at "
              f"{a.assume_lambda}. This curve\n    shows only how Cp_elec "
              f"would vary with wind speed at fixed λ — it cannot\n    "
              f"locate a peak in λ, which is the entire purpose of Cp(λ).")

    if a.csv:
        out = Path(a.csv)
        out.parent.mkdir(parents=True, exist_ok=True)
        with out.open("w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["# rotor_speed_source", src])
            w.writerow(["# rotor_type", a.rotor])
            w.writerow(["# radius_m", a.radius]); w.writerow(["# height_m", a.height])
            w.writerow(["# rho", rho]); w.writerow(["# area_m2", A])
            w.writerow(["# quantity", "Cp_elec = Cp_aero * eta_gen * eta_rect"])
            w.writerow(["mps", "fan_rpm", "rotor_rpm", "lambda", "p_w",
                        "avail_w", "cp_elec"])
            for r in res:
                w.writerow([f"{r['mps']:.3f}", f"{r.get('fan_rpm',0):.0f}",
                            f"{r['rpm']:.1f}", f"{r['lam']:.4f}",
                            f"{r['p_w']:.5f}", f"{r['avail_w']:.4f}",
                            f"{r['cp_elec']:.6f}"])
        print(f"\n  wrote {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
